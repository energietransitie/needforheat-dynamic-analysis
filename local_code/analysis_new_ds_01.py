import tkinter

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from gekko import GEKKO
import openpyxl
import os

########################################################################################################################
#                                                     Pandas Setting
########################################################################################################################
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

########################################################################################################################
#                                                     Load Data
########################################################################################################################
file_name = '../dataset/assendorp/simulatedhome_Henri.xlsx'
# df = pd.read_excel(file_name, sheet_name='simulated home')
df = pd.read_excel(file_name, sheet_name='virtual home')
df_input = pd.read_excel(file_name, sheet_name='input')


flat_pseudonym = df['home_id'].tolist()[0]

########################################################################################################################
#                                                   Data Pre-Process
########################################################################################################################
delta_t = 900
end_point = df.index[df['timestamp'] == '2022-01-03 01:00:00'].tolist()[0]
sim_start_point = end_point
# printing_row = 2
duration = 12
num_iter = 1

########################################################################################################################
#                                                   Gekko Core
########################################################################################################################

for i in range(num_iter):
    start_point = end_point
    end_point = start_point + (duration * 24 * 4) - 1

    # T_in_meas = np.asarray(df['T_in_avg_C'].iloc[start_point:end_point])  # [K]
    T_in_meas = np.asarray(df_input['T_in_avg_C'].iloc[start_point:end_point])  # [K]

    T_out = np.asarray(df['T_out_avg_C'].iloc[start_point:end_point])  # [K]
    T_out_eff_arr = np.asarray(df['T_out_e_avg_C'].iloc[start_point:end_point])  # [K]

    I_geo_eff_val = np.asarray(df['irradiation_hor_avg_W_per_m2'].iloc[start_point:end_point])  # [W/m2]
    delta_G_val = np.asarray(df['gas_m3_per_interval'].iloc[start_point:end_point])  # [m^3]
    delta_E_int_val = np.asarray(df['e_remaining_heat_kWh_per_interval'].iloc[start_point:end_point])  # [kWh]

    # delta_Q_gain_arr = np.asarray(df['Q_gain_avg_W'].iloc[start_point:end_point])             # [W]

    delta_Q_gain_gas_no_CH_avg_W = np.asarray(df['Q_gain_gas_no_CH_avg_W'].iloc[start_point:end_point])     #[W]
    delta_Q_gain_int_occup_avg_W = np.asarray(df['Q_gain_int_occup_avg_W'].iloc[start_point:end_point])     #[W]
    delta_e_remaining_heat_avg_W = np.asarray(df['e_remaining_heat_avg_W'].iloc[start_point:end_point])     #[W]
    delta_Q_gain_int_avg_W_arr = delta_Q_gain_gas_no_CH_avg_W + delta_Q_gain_int_occup_avg_W + delta_e_remaining_heat_avg_W    #[W]

    irradiation_hor_avg_W_per_m2_arr = np.asarray(df['irradiation_hor_avg_W_per_m2'].iloc[start_point:end_point])  #[W/m2]

    heating_frac_arr = np.asarray(df['heating_frac'].iloc[start_point:end_point])

    # delta_Q_gain_int_avg_W = np.asarray(df['Q_gain_int_avg_W'].iloc[start_point:end_point])     # [W]
    # delta_q_gain_sol_avg_W_arr = np.asarray(df['Q_gain_sol_avg_W'].iloc[start_point:end_point])     # [W]
    # delta_Q_gain_CH_avg_W_arr = np.asarray(df['Q_gain_CH_avg_W'].iloc[start_point:end_point])       # [W]
    # delta_Q_gain_arr = delta_Q_gain_int_avg_W + delta_q_gain_sol_avg_W + delta_Q_gain_CH_avg_W  # [W]
    # delta_Q_gain_arr = delta_Q_gain_int_avg_W + delta_Q_gain_CH_avg_W  # [W]

    # delta_Q_loss_arr = np.asarray(df['Q_loss_avg_W'].iloc[start_point:end_point])  # [W]

    # # data loading
    # e_used_normal_val = np.asarray(df['e_used_normal_kWh'].iloc[start_point:end_point]) # [kWh]
    # e_used_low_val = np.asarray(df['e_used_low_kWh'].iloc[start_point:end_point]) # [kWh]
    # delta_E_supply_val = np.asarray(e_used_normal_val + e_used_low_val) # [kWh]
    #
    # e_returned_normal_val = np.asarray(df['e_returned_normal_kWh'].iloc[start_point:end_point]) # [kWh]
    # e_returned_low_val = np.asarray(df['e_returned_low_kWh'].iloc[start_point:end_point]) # [kWh]
    # delta_E_ret_val = np.asarray(e_returned_normal_val + e_returned_low_val) # [kWh]
    # delta_E_int_val = delta_E_supply_val - delta_E_ret_val  # [kWh]

    #
    time_recorded = df['timestamp'].iloc[start_point:end_point] # [s]
    #
    #
    # delta_G_noCH_val = 339.0 / (365.25 * 24 * 60 * 60)
    # delta_G_CH_val = delta_G_val - delta_G_noCH_val
    # delta_G_CH_val[delta_G_CH_val < 0] = 0

    m = GEKKO(remote=False)
    # m.time = np.linspace(delta_t, len(T_in_meas) * delta_t, len(T_in_meas))  # [s]
    m.time = np.linspace(0, (len(T_in_meas)-1) * delta_t, len(T_in_meas))  # [s]

    # Gekko constant Parameter
    # h_sup = m.Param(value=35170000) # [J/Nm^3]
    # COP = m.Param(value=6)
    # h_E = m.Param(value=60 * 60 * 1000) # [J/kWh]
    eta_sup_CH = m.Param(value=0.97)
    # eta_no_ch = m.Param(value=0.34)
    # delta_G_noCH = m.Param(value=339.0 / (365.25 * 24 * 60 * 60)) #[Nm^3/s]
    # delta_E_CH = m.Param(value=0)
    # delta_Q_int_occup = m.Param(value=2.2 * 61 / (24*3600)) #[W]
    A_eff = m.Param(value=12)                   #[m^2]
    P_kw = m.Param(value=15)                    #[kW]

    # Gekko Fixed Variable (model parameter)
    ## tau Input: the following value should be based on hour for tau [hr]
    tau_init_val_hr = 100
    tau_lb_hr = 10
    tau_ub_hr = 1000

    ## Internal conversion (do not change this part)
    tau_init_val = tau_init_val_hr * 3600
    tau_lb = tau_lb_hr * 3600
    tau_ub = tau_ub_hr * 3600

    # A_eff = m.FV(value=6) ; A_eff.STATUS=1 ; A_eff.FSTATUS=0 #[m^2]
    tau = m.FV(value=tau_init_val, lb=tau_lb, ub=tau_ub); tau.STATUS = 1; tau.FSTATUS = 0  # [s]
    H = m.FV(value=300.0, lb=0, ub=1000); H.STATUS = 1; H.FSTATUS = 0  # [W/K]

    # Gekko Manipulated Variable
    # delta_G = m.MV(value=gas_total / delta_t) ; delta_G.STATUS=0 ; delta_G.FSTATUS=1 # [Nm^3/s]
    # delta_G_CH = m.MV(value=delta_G_CH_val) ; delta_G_CH.STATUS=0 ; delta_G_CH.FSTATUS=1 # [Nm^3/s]
    # I_geo = m.MV(value=I_geo_eff_val) ; I_geo.STATUS=0 ; I_geo.FSTATUS=1 #[W/m^2]
    # delta_E_int = m.MV(value=delta_E_int_val) ; delta_E_int.STATUS=0 ; delta_E_int.FSTATUS=1 #[kWh]
    T_out_eff = m.MV(value=T_out_eff_arr); T_out_eff.STATUS = 0; T_out_eff.FSTATUS = 1 #[K]

    # Gekko Control Variable
    T_in_sim = m.CV(value=T_in_meas); T_in_sim.STATUS=1 ; T_in_sim.FSTATUS=1 #[deg C]

    # Control Variable options:
    # T_in_sim.MEAS_GAP= 0.1
    # T_in_sim.WMEAS = 0.1

    # delta_Q_CH
    # delta_G_CH = m.Intermediate(delta_G - delta_G_noCH) #[w]
    # delta_Q_CH = m.Intermediate((delta_G_CH * eta_hs * h_sup) + (delta_E_CH * COP * h_E)) #[w]

    # delta_Q_sol
    # delta_Q_sol = m.Intermediate(A_eff * I_geo/delta_t) #[w]
    # delta_Q_sol = m.Intermediate(A_eff * I_geo) #[w]

    # delta_Q_int
    # delta_Q_int_e = m.Intermediate(delta_E_int * h_E / delta_t)    #[w]
    # delta_Q_int_gas_noCH = m.Intermediate(delta_G_noCH * eta_no_ch * h_sup)         #[w]
    # delta_Q_int = m.Intermediate(delta_Q_int_e + delta_Q_int_occup + delta_Q_int_gas_noCH)      #[w]
    # delta_Q_int_other = m.Intermediate(delta_Q_int_e + delta_Q_int_occup)

    # delta_Q_gain
    # delta_Q_gain = m.Intermediate(delta_Q_CH + delta_Q_sol + delta_Q_int) #[w]

    # C_eff = m.Intermediate(H * tau) #[J/K]
    # m.Equation(T_in_sim.dt() == ((delta_Q_gain - (H * (T_in_sim - T_out_eff))) * delta_t / C_eff ))  #


    delta_Q_gain_int_avg_W = m.MV(value=delta_Q_gain_int_avg_W_arr)
    delta_Q_gain_int_avg_W.STATUS = 0
    delta_Q_gain_int_avg_W.FSTATUS = 1

    irradiation_hor_avg_W_per_m2 = m.MV(value=irradiation_hor_avg_W_per_m2_arr)
    irradiation_hor_avg_W_per_m2.STATUS = 0
    irradiation_hor_avg_W_per_m2.FSTATUS = 1
    delta_q_gain_sol_avg_W = m.Intermediate(A_eff * irradiation_hor_avg_W_per_m2)                   # [W]

    heating_frac = m.MV(value=heating_frac_arr)
    heating_frac.STATUS = 0
    heating_frac.FSTATUS = 1
    delta_Q_gain_CH_avg_W = m.Intermediate(P_kw * 1000 * eta_sup_CH * heating_frac)

    delta_Q_gain = m.Intermediate(delta_Q_gain_int_avg_W + delta_q_gain_sol_avg_W + delta_Q_gain_CH_avg_W)  # [W]

    m.Equation(T_in_sim.dt() == (delta_Q_gain - (H * (T_in_sim - T_out_eff))) / (H * tau))

########################################################################################################################
#                                                    Solve Equations
########################################################################################################################
    m.options.IMODE = 5
    m.options.EV_TYPE = 2  # specific objective function (L1-norm vs L2-norm)
    m.options.NODES = 2
    # m.options.SOLVER = 3  # IPOPT

    # add dead-band for measurement to avoid over-fitting
    m.solve(disp=False)

    print(
        "Start point: {} ====> End point: {}".format(df.loc[start_point, 'timestamp'], df.loc[end_point, 'timestamp']))
    print('Iter: ', i)
    print('effective thermal inertia: tau [hr]: ' + str(round(tau.value[0] / 3600, 2)))
    print('specific heat loss: H [W/K]: ' + str(round(H.value[0], 2)))
    print('Effective area of solar aperture in the horizontal plane: A_eff [m^2]: ' + str(A_eff.value[0]))
    # print('internal device heat gain: ' + str(delta_Q_int.value[0]))
    obj = m.options.OBJFCNVAL
    print('value obj: ', obj)
    print('+' * 60)

    temp_diff = []
    for j, k in zip(T_in_sim, T_in_meas):
        temp_diff.append(abs(j - k))

    print('temp_diff: ', temp_diff)
    print('+' * 60)
    print('diff sum: ', sum(temp_diff))
    print('+' * 60)
    print('diff avg: ', sum(temp_diff) / len(temp_diff))
    print('*' * 50)
    print('max temp_diff', max(temp_diff))
    print('len T_in_sim: ', len(T_in_sim))
    print('sim_start point: {} and end_point: {}'.format(sim_start_point, end_point))
    df.loc[start_point:end_point - 1, ['T_in_sim']] = list(T_in_sim.value)

    duration_s = m.time[-1]
    error_K_total_time = (m.options.OBJFCNVAL ** (1 / m.options.EV_TYPE)) / duration_s
    # error_K_15_minutes = (m.options.OBJFCNVAL ** (1 / m.options.EV_TYPE)) / delta_t
    # error_K_data_point = (m.options.OBJFCNVAL ** (1 / m.options.EV_TYPE)) / (7*24*4)

    # row_num = printing_row + i
    # result_excel = '../Analysis/assendorp/result/A_fixed/result_new_dataset.xlsx'
    # wb = openpyxl.load_workbook(filename=result_excel)
    # ws = wb.worksheets[0]
    # ws.cell(row=row_num, column=2).value = flat_pseudonym
    # ws.cell(row=row_num, column=3).value = df['timestamp'][start_point]
    # ws.cell(row=row_num, column=4).value = df['timestamp'][end_point]
    # ws.cell(row=row_num, column=5).value = duration * (24 * 60 * 60)
    # ws.cell(row=row_num, column=6).value = error_K_total_time
    # ws.cell(row=row_num, column=7).value = round(H.value[0], 2)
    # ws.cell(row=row_num, column=8).value = round(tau.value[0] / 3600, 2)
    # # ws.cell(row=row_num, column=9).value = eta_hs
    # ws.cell(row=row_num, column=10).value = A_eff.value[0]
    # wb.save(result_excel)


    # plt.subplot(2,1,1)
    plt.plot(time_recorded, T_in_meas, 'r*', label='Measured')
    plt.plot(time_recorded, T_in_sim.value, 'bo', label='Simulated')
    plt.plot(time_recorded, T_out_eff_arr, 'gx', label='outdoor temp')
    plt.xlabel('Date and Time', fontsize=14)
    plt.ylabel('Temperature ($^\circ$C)', fontsize=13)
    plt.legend()

    # plt.subplot(2,1,2)
    # plt.plot(time_recorded, delta_Q_sol, 'b*' , label='solar irradiation [w]')
    # plt.plot(time_recorded, delta_Q_CH, 'g-', label = 'delta_Q_CH [w]')
    # plt.plot(time_recorded, delta_Q_int_gas_noCH, 'ro', label='delta_Q_int_gas_noCH [w]')
    # plt.plot(time_recorded, delta_Q_int_occup, 'k+', label='delta_Q_int_occup [w]')
    # plt.plot(time_recorded, delta_Q_int_e, '--', label='delta_Q_int_e [w]')
    # plt.xlabel('Date and Time', fontsize=14)
    # plt.ylabel('energy [w]', fontsize=13)
    # plt.legend()
    plt.show()